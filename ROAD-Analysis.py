import os
import sqlite3
import pandas as pd
import openpyxl
import argparse
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from utils.user_queries import USER_QUERIES
from utils.tenant_queries import TENANT_QUERIES

def print_banner():
    ascii_art = """
██▀███  ▒█████  ▄▄▄     ▓█████▄▄▄▄█████▓▒█████  ▒█████  ██▓     ██████     ▄▄▄      ███▄    █ ▄▄▄      ██▓  ▓██   ██▓ ██████ ██▓ ██████ 
▓██ ▒ ██▒██▒  ██▒████▄   ▒██▀ ██▓  ██▒ ▓▒██▒  ██▒██▒  ██▓██▒   ▒██    ▒    ▒████▄    ██ ▀█   █▒████▄   ▓██▒   ▒██  ██▒██    ▒▓██▒██    ▒ 
▓██ ░▄█ ▒██░  ██▒██  ▀█▄ ░██   █▒ ▓██░ ▒▒██░  ██▒██░  ██▒██░   ░ ▓██▄      ▒██  ▀█▄ ▓██  ▀█ ██▒██  ▀█▄ ▒██░    ▒██ ██░ ▓██▄  ▒██░ ▓██▄   
▒██▀▀█▄ ▒██   ██░██▄▄▄▄██░▓█▄   ░ ▓██▓ ░▒██   ██▒██   ██▒██░     ▒   ██▒   ░██▄▄▄▄██▓██▒  ▐▌██░██▄▄▄▄██▒██░    ░ ▐██▓░ ▒   ██░██░ ▒   ██▒
░██▓ ▒██░ ████▓▒░▓█   ▓██░▒████▓  ▒██▒ ░░ ████▓▒░ ████▓▒░██████▒██████▒▒    ▓█   ▓██▒██░   ▓██░▓█   ▓██░██████▒░ ██▒▓▒██████▒░██▒██████▒▒
░ ▒▓ ░▒▓░ ▒░▒░▒░ ▒▒   ▓▒█░▒▒▓  ▒  ▒ ░░  ░ ▒░▒░▒░░ ▒░▒░▒░░ ▒░▓  ▒ ▒▓▒ ▒ ░    ▒▒   ▓▒█░ ▒░   ▒ ▒ ▒▒   ▓▒█░ ▒░▓  ░ ██▒▒▒▒ ▒▓▒ ▒ ░▓ ▒ ▒▓▒ ▒ ░
░▒ ░ ▒░ ░ ▒ ▒░  ▒   ▒▒ ░░ ▒  ▒    ░     ░ ▒ ▒░  ░ ▒ ▒░░ ░ ▒  ░ ░▒  ░ ░     ▒   ▒▒ ░ ░░   ░ ▒░ ▒   ▒▒ ░ ░ ▒  ▓██ ░▒░░ ░▒  ░ ░▒ ░ ░▒  ░ ░
░░   ░░ ░ ░ ▒   ░   ▒   ░ ░  ░  ░     ░ ░ ░ ▒ ░ ░ ░ ▒   ░ ░  ░  ░  ░       ░   ▒     ░   ░ ░  ░   ▒    ░ ░  ▒ ▒ ░░ ░  ░  ░  ▒ ░  ░  ░  
░        ░ ░       ░  ░  ░               ░ ░     ░ ░     ░  ░     ░           ░  ░        ░      ░  ░   ░  ░ ░          ░  ░       ░  
By @FlyingPhishy
"""
    print(ascii_art)

def connect_to_db(db_file):
    try:
        return sqlite3.connect(db_file)
    except sqlite3.Error as e:
        print(f"Database connection error: {e}")
        return None

def run_query(conn, query):
    try:
        return pd.read_sql_query(query, conn)
    except sqlite3.Error as e:
        print(f"Error running query: {e}")
        return pd.DataFrame()

def is_json_string(x):
    if not isinstance(x, str):
        return False
    x = x.strip()
    return (x.startswith('{') and x.endswith('}')) or (x.startswith('[') and x.endswith(']'))

def format_json(value):
    try:
        # Parse the JSON string
        parsed = json.loads(value)
        # Format it with indentation
        return json.dumps(parsed, indent=2)
    except json.JSONDecodeError:
        # If it's not valid JSON, return the original value
        return value

def save_to_excel(dataframes, file_name):
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            # Ensure the DataFrame has at least one row (for headers)
            if df.empty:
                df = pd.DataFrame(columns=df.columns)
                df.loc[0] = [''] * len(df.columns)  # Add an empty row
            
            # Identify columns that contain JSON
            json_columns = df.columns[df.apply(lambda col: col.map(is_json_string).any())]
            
            # Format JSON in identified columns
            for col in json_columns:
                df[col] = df[col].map(format_json)
            
            # Write the dataframe to Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the worksheet
            worksheet = writer.sheets[sheet_name]
            
            for idx, col in enumerate(df.columns):
                column_letter = get_column_letter(idx + 1)
                
                # Set a reasonable maximum width
                max_width = 100
                
                if col in json_columns:
                    # For JSON columns, set a fixed width and wrap text
                    worksheet.column_dimensions[column_letter].width = max_width
                    for cell in worksheet[column_letter]:
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                else:
                    # For other columns, auto-fit width
                    max_length = max(df[col].astype(str).map(len).max(), len(str(col)))
                    adjusted_width = min(max_length + 2, max_width)  # +2 for some padding
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze the header row
            worksheet.freeze_panes = 'A2'
            
            # If the DataFrame was originally empty, remove the empty data row
            if df.shape[0] == 1 and all(df.iloc[0].isna()):
                for row in worksheet['A2:' + get_column_letter(worksheet.max_column) + '2']:
                    for cell in row:
                        cell.value = None

    print(f"Excel file saved: {file_name}")

def print_summary_stats(stats_df):
    """Print a summary of key statistics to the console."""
    stats = stats_df.iloc[0]  # Get the first (and only) row of the DataFrame
    
    print("\nGeneral Statistics:")
    print(f"* Total Users: {stats['Total Users']}")
    print(f"* Total Member Users: {stats['Total Member Users']}")
    print(f"* Total Active Users: {stats['Total Active Users']}")
    print(f"* Total Guests: {stats['Total Guests']}")
    print(f"* Total Active Guest Users: {stats['Total Active Guests']}")
    print(f"* Total Users w/ Disable Password Expiry: {stats['Users w/ Disable Password Expiry']}")
    
    print("\nPassword statistics for active users:")
    print(f"* {stats['Percentage Users w/ Disable Password Expiry']:.1f}% of users have the password policy 'Disable Password Expiration' attached.")
    print(f"* {stats['Percentage of Members Password > 90 Days']:.1f}% of member users have passwords over 90 days old.")
    print(f"* {stats['Percentage of Guests Password > 90 Days']:.1f}% of guest users have passwords over 90 days old.")
    print(f"* {stats['Percentage of Members w/ Unchanged Passwords From Creation']:.1f}% of members have not changed their password since creation.")
    print(f"* {stats['Percentage of Guests w/ Unchanged Passwords From Creation']:.1f}% of guest users have not changed their password since creation.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run SQLite queries and export results to Excel.")
    parser.add_argument("-db", "--database", required=True, help="Path to the SQLite database file.")
    parser.add_argument("-o", "--output", required=True, help="Path and name of the output Excel file.")
    args = parser.parse_args()

    print_banner()

    if not os.path.exists(args.database):
        print(f"The specified database file does not exist: {args.database}")
        exit()

    database_file = args.database
    output_file = args.output

    conn = connect_to_db(database_file)
    if conn:
        user_results = {}
        tenant_results = {}
        
        # Run user queries
        for name, query in USER_QUERIES.items():
            user_results[name] = run_query(conn, query)
        
        # Run tenant queries
        for name, query in TENANT_QUERIES.items():
            tenant_results[name] = run_query(conn, query)
        
        # Combine results for saving to Excel
        all_results = {f"User_{k}": v for k, v in user_results.items()}
        all_results.update({f"Tenant_{k}": v for k, v in tenant_results.items()})
        
        save_to_excel(all_results, output_file)
        
        # Print summary statistics to console
        if 'OverallStats' in user_results:
            print_summary_stats(user_results['OverallStats'])
        
        conn.close()